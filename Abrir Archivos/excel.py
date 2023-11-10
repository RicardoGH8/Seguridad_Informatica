import openpyxl

def contenido_excel(nombre_archivo):
    try:
        workbook = openpyxl.load_workbook(nombre_archivo)

        sheet = workbook.active

        print(f"Contenido del archivo '{nombre_archivo}':")
        
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            row_data = [cell.value for cell in row]
            print(row_data)

    except Exception as e:
        print(f"Error al abrir el archivo Excel: {e}")

archivo_excel = 'excel.xlsx'
contenido_excel(archivo_excel)
